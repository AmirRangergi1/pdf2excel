from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pathlib import Path
import os

from backend.services.openai_client import OpenAIClient
from backend.services.processor import parse_model_response

BASE_DIR = Path(__file__).resolve().parents[1]
FRONTEND_DIST = BASE_DIR / "frontend" / "dist"
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", str(BASE_DIR.parent / "uploads")))
PUBLIC_BASE = os.getenv("PUBLIC_BASE_URL")  # optional, e.g. https://your.domain

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI()

if FRONTEND_DIST.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIST), html=True), name="static")
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

openai_client = OpenAIClient()


@app.get("/api/health")
async def health():
    return {"status": "ok"}


@app.post("/api/convert")
async def convert(file: UploadFile | None = File(None), file_url: str | None = Form(None), question: str | None = Form(None)):
    if file is None and not file_url:
        raise HTTPException(status_code=400, detail="Provide uploaded file or file_url")

    if file is not None:
        dest = UPLOAD_DIR / file.filename
        with dest.open("wb") as f:
            f.write(await file.read())
        if PUBLIC_BASE:
            file_url = PUBLIC_BASE.rstrip("/") + "/uploads/" + file.filename
        else:
            raise HTTPException(status_code=400, detail="Uploaded files require PUBLIC_BASE_URL to generate a public URL for the model to fetch")

    prompt = question or "Convert the provided PDF to an Excel file and return it as base64 in JSON: {\"filename\":..., \"file_base64\":... }"

    resp = openai_client.create_chat_completion_with_file(file_url=file_url, question=prompt)

    # parse model response to get filename and bytes
    parsed = parse_model_response(resp)
    if not parsed:
        return JSONResponse(status_code=500, content={"error": "Could not parse model response"})

    out_name, out_bytes = parsed
    out_path = UPLOAD_DIR / out_name
    with out_path.open("wb") as f:
        f.write(out_bytes)

    return FileResponse(path=str(out_path), filename=out_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("backend.main:app", host="0.0.0.0", port=int(os.getenv("PORT", "8000")), log_level="info")
import io
import os
import base64
import logging
from pathlib import Path

import fitz  # PyMuPDF
import requests
import openpyxl
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

# Basic logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("pdf2excel")

# Config from environment (user-provided defaults may be set here)
LIARA_API_URL = os.environ.get("LIARA_API_URL", "https://ai.liara.ir/api/695a383ba042f871d25e6074/v1")
LIARA_API_KEY = os.environ.get("LIARA_API_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJrZXkiOiI2OTVlZjBjNWJjN2I4NzkzNmJjNWQyODQiLCJ0eXBlIjoiYWlfa2V5IiwiaWF0IjoxNzY3ODI5NzAxfQ.Pn1GEoWGZluG87n81WKT2WsBMsy5Ov8NCbrgWt1xqWA")
LIARA_MODEL_ID = os.environ.get("LIARA_MODEL_ID", "anthropic/claude-sonnet-4.5")

# Feature availability
try:
    import fitz  # noqa: F401
    PYMUPDF_AVAILABLE = True
except Exception:
    PYMUPDF_AVAILABLE = False

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve frontend if built (mount at /static to avoid catching API routes)
frontend_dist = Path(__file__).parent.parent / "frontend" / "dist"
if frontend_dist.exists():
    app.mount("/static", StaticFiles(directory=str(frontend_dist), html=True), name="static")
    # Also expose the assets directory at /assets so built files can be fetched
    assets_dir = frontend_dist / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=str(assets_dir), html=False), name="assets")


def pdf_to_images_base64(pdf_bytes: bytes, zoom: float = 2.0) -> list:
    """Render PDF pages to PNG images (base64) using PyMuPDF.

    Returns a list of dicts: {"page": n, "image": <base64 png>}.
    """
    images = []
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            png = pix.tobytes("png")
            b64 = base64.b64encode(png).decode("utf-8")
            images.append({"page": page_num + 1, "image": b64})
        doc.close()
    except Exception as e:
        logger.error(f"Error converting PDF to images: {e}")
    return images


def extract_tables_with_liara(pdf_content: bytes, filename: str) -> list:
    """Send PDF pages (images) to Liara AI endpoint and parse table JSON.

    Expected request shape: {"filename": ..., "pages": [{"page":1, "image": "...base64..."}, ...]}
    The response is flexible; this function handles a few common shapes.
    """
    if not LIARA_API_URL or not LIARA_API_KEY:
        raise ValueError("Liara API not configured (LIARA_API_URL/LIARA_API_KEY)")

    images = pdf_to_images_base64(pdf_content)
    if not images:
        raise ValueError("Unable to render PDF pages to images")

    payload = {"filename": filename, "model": LIARA_MODEL_ID, "pages": images}
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {LIARA_API_KEY}",
    }

    try:
        logger.info(f"Sending {len(images)} pages to Liara: {LIARA_API_URL}")
        resp = requests.post(LIARA_API_URL, json=payload, headers=headers, timeout=120)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        logger.error(f"Liara request failed: {e}")
        raise

    tables_out = []

    # Top-level tables
    if isinstance(data, dict) and data.get("tables"):
        for t in data.get("tables", []):
            tables_out.append({"page": t.get("page", 1), "data": t, "type": "liara"})
        if tables_out:
            return tables_out

    # Per-page tables
    if isinstance(data, dict) and data.get("pages"):
        for pg in data.get("pages", []):
            pg_idx = pg.get("page", 1)
            for t in pg.get("tables", []) or []:
                tables_out.append({"page": pg_idx, "data": t, "type": "liara"})
        if tables_out:
            return tables_out

    # If response is a list of tables
    if isinstance(data, list):
        for t in data:
            if isinstance(t, dict):
                tables_out.append({"page": t.get("page", 1), "data": t, "type": "liara"})
        if tables_out:
            return tables_out

    # Fallback: if Liara returned text, place it into a single-sheet result
    logger.warning("Liara returned unexpected shape; creating fallback sheet")
    fallback = {"headers": ["Liara Output"], "rows": [[str(data)[:100]]]} if data else {"headers": ["No Data"], "rows": [["No tables extracted"]]}
    return [{"page": 1, "data": fallback, "type": "fallback"}]


def create_excel_from_tables(tables: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    # Remove default sheet
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    for idx, tinfo in enumerate(tables, start=1):
        page = tinfo.get("page", 1)
        data = tinfo.get("data", {}) or {}
        sheet_name = f"P{page}_{idx}"[:31]
        ws = wb.create_sheet(sheet_name)

        headers = data.get("headers") or data.get("columns") or []
        rows = data.get("rows") or data.get("data") or []

        if headers:
            for c, h in enumerate(headers, start=1):
                ws.cell(row=1, column=c, value=h)
        # If rows are list of dicts, normalize
        if rows and isinstance(rows[0], dict):
            # write headers as union of keys if none provided
            if not headers:
                headers = list(rows[0].keys())
                for c, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=c, value=h)
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, h in enumerate(headers, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=row.get(h))
        else:
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

        # adjust widths
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.post("/api/convert")
async def convert_pdf_to_excel(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")

    try:
        content = await file.read()
        tables = extract_tables_with_liara(content, file.filename)
        excel = create_excel_from_tables(tables)

        return StreamingResponse(
            iter([excel.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{file.filename.replace('.pdf','')}.xlsx\""},
        )
    except ValueError as e:
        logger.error(f"Conversion error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("Unexpected error during conversion")
        # return an Excel with the error message
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Conversion Error"
        ws["A2"] = str(e)[:100]
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"error_{file.filename.replace('.pdf','')}.xlsx\""},
        )


@app.get("/api/health")
async def health():
    return {
        "status": "healthy",
        "service": "PDF to Excel Converter",
        "pymupdf_available": PYMUPDF_AVAILABLE,
        "liara_configured": bool(LIARA_API_URL and LIARA_API_KEY),
        "liara_url": LIARA_API_URL if LIARA_API_URL else None,
    }


@app.get("/")
async def root():
    try:
        index_file = frontend_dist / "index.html"
        if index_file.exists():
            return FileResponse(str(index_file), media_type="text/html")
    except Exception:
        pass
    return {
        "name": "PDF to Excel Converter",
        "version": "2.0",
        "powered_by": "Liara AI",
        "endpoints": {"convert": "/api/convert (POST)", "health": "/api/health (GET)"},
    }


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
